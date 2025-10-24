import csv
from openpyxl import load_workbook

# Đọc dữ liệu từ file CSV, trả về list dict
def read_csv(filepath):
    with open(filepath, encoding="utf-8") as f:
        reader = csv.DictReader(f)
        return [row for row in reader]

# Đọc dữ liệu từ file Excel, trả về list dict
# Mỗi sheet là một list dict, key là tên cột
def read_excel(filepath, sheet_name=None):
    wb = load_workbook(filepath, data_only=True)
    sheets = [sheet_name] if sheet_name else wb.sheetnames
    data = {}
    for sheet in sheets:
        ws = wb[sheet]
        headers = [cell.value for cell in ws[1]]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append(dict(zip(headers, row)))
        data[sheet] = rows
    return data if not sheet_name else data[sheet_name]
